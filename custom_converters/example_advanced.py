"""
Advanced Custom Converter - Autocrat Attendance v0.0.1
Demonstrates full control over output format.
"""

import pandas as pd
from datetime import datetime

def convert(input_path, output_path, config=None):
    print(f"[Advanced] Input: {input_path}")
    print(f"[Advanced] Output: {output_path}")
    print(f"[Advanced] Config: {config}")
    xls = pd.ExcelFile(input_path)
    dfs = []
    for sheet in xls.sheet_names:
        try:
            temp = pd.read_excel(input_path, sheet_name=sheet)
            dfs.append(temp)
        except:
            continue
    raw = pd.concat(dfs, ignore_index=True) if len(dfs)>1 else dfs[0]
    emp_col = None
    for col in raw.columns:
        if 'emp' in str(col).lower() and 'id' in str(col).lower():
            emp_col = col
            break
    if not emp_col:
        emp_col = raw.columns[0]
    grouped = raw.groupby(emp_col).size().reset_index(name='Punch Count')
    grouped['Processed At'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    grouped['Company'] = config.get('company_name', 'Autocrat Solutions') if config else 'Autocrat Solutions'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        raw.to_excel(writer, sheet_name='Raw Data', index=False)
        grouped.to_excel(writer, sheet_name='Employee Punch Count', index=False)
        cfg = pd.DataFrame([
            {'Parameter': 'Input File', 'Value': input_path},
            {'Parameter': 'Output File', 'Value': output_path},
            {'Parameter': 'Converter', 'Value': 'example_advanced.py'},
            {'Parameter': 'Generated', 'Value': datetime.now().isoformat()},
            {'Parameter': 'Total Employees', 'Value': grouped.shape[0]},
        ])
        cfg.to_excel(writer, sheet_name='Config', index=False)
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col in range(1, ws.max_column+1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
    print(f"[Advanced] Done -> {output_path}")
    return {
        "success": True,
        "output_path": output_path,
        "summary": grouped.to_dict(orient='records'),
        "total_employees": len(grouped)
    }

process = convert
process_attendance = convert
main = convert
