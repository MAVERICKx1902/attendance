"""
Autocrat Solutions - Attendance Processor Core
Liquid Glass Edition

This module converts raw biometric / manual attendance Excel dumps into
formatted, HR-ready monthly attendance sheets.

Input: Raw Excel (.xlsx/.xls) with at least:
 - Employee ID / Code
 - Employee Name (optional)
 - Date
 - Time / Punch Time / DateTime
 - Optional: Department, Shift, Device

Output: Formatted Excel with 4 sheets:
 1. Cleaned Logs
 2. Daily Matrix (P/A/L/H/HD + Late + Hours)
 3. Monthly Summary
 4. Config & Holidays

Works standalone (CLI) or imported by app.py / frontend via pyodide.
"""

import re
from datetime import datetime, timedelta, time
from typing import Dict, List, Tuple, Optional
import pandas as pd
import numpy as np

# Column mapping heuristics
COLUMN_ALIASES = {
    'emp_id': ['emp id', 'employee id', 'employee code', 'emp code', 'id', 'code', 'staff id', 'card no', 'enroll', 'user id'],
    'emp_name': ['emp name', 'employee name', 'name', 'staff name', 'employee', 'person'],
    'datetime': ['datetime', 'date time', 'punch datetime', 'timestamp', 'log time'],
    'date': ['date', 'punch date', 'attendance date', 'day'],
    'time': ['time', 'punch time', 'log time', 'in time', 'out time', 'check time'],
    'department': ['department', 'dept', 'division', 'team'],
    'status': ['status', 'in/out', 'io', 'type', 'punch', 'direction'],
}

def normalize_col(col: str) -> str:
    return re.sub(r'[^a-z0-9 ]', '', str(col).lower().strip())

def detect_columns(df: pd.DataFrame) -> Dict[str, str]:
    norm_map = {normalize_col(c): c for c in df.columns}
    detected = {}
    for logical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_col(alias)
            # exact or partial
            for ncol, orig in norm_map.items():
                if alias_norm == ncol or alias_norm in ncol or ncol in alias_norm:
                    if logical not in detected:
                        detected[logical] = orig
                        break
            if logical in detected:
                break
    return detected

def parse_datetime_series(date_series, time_series=None, datetime_series=None):
    """Robust datetime parsing"""
    if datetime_series is not None and not datetime_series.isna().all():
        parsed = pd.to_datetime(datetime_series, errors='coerce', dayfirst=True)
        if parsed.notna().sum() > 0:
            return parsed

    if date_series is not None:
        dates = pd.to_datetime(date_series, errors='coerce', dayfirst=True)
        if time_series is not None and not time_series.isna().all():
            times = pd.to_datetime(time_series, errors='coerce').dt.time
            # combine
            combined = []
            for d, t in zip(dates, times):
                if pd.isna(d):
                    combined.append(pd.NaT)
                else:
                    if pd.isna(t):
                        combined.append(d)
                    else:
                        try:
                            combined.append(datetime.combine(d.date(), t))
                        except:
                            combined.append(d)
            return pd.Series(combined)
        return dates
    if time_series is not None:
        return pd.to_datetime(time_series, errors='coerce', dayfirst=True)
    return pd.Series([pd.NaT]*len(date_series if date_series is not None else []))

def clean_dataframe(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    detected = detect_columns(df)
    # print for debug
    # Try to construct unified columns
    emp_id_col = detected.get('emp_id')
    emp_name_col = detected.get('emp_name')
    datetime_col = detected.get('datetime')
    date_col = detected.get('date')
    time_col = detected.get('time')
    dept_col = detected.get('department')

    # Build working frame
    work = pd.DataFrame()
    if emp_id_col:
        work['emp_id'] = df[emp_id_col].astype(str).str.strip()
    else:
        work['emp_id'] = df.iloc[:,0].astype(str)  # fallback first column

    if emp_name_col:
        work['emp_name'] = df[emp_name_col].astype(str).str.strip()
    else:
        work['emp_name'] = work['emp_id']  # fallback

    if dept_col:
        work['department'] = df[dept_col].astype(str)
    else:
        work['department'] = 'General'

    # datetime parsing
    date_s = df[date_col] if date_col in df.columns else None
    time_s = df[time_col] if time_col in df.columns else None
    dt_s = df[datetime_col] if datetime_col in df.columns else None

    work['punch_datetime'] = parse_datetime_series(date_s, time_s, dt_s)
    work = work.dropna(subset=['punch_datetime'])
    work['punch_datetime'] = pd.to_datetime(work['punch_datetime'])
    work['date'] = work['punch_datetime'].dt.date
    work['time'] = work['punch_datetime'].dt.time
    work['date_str'] = work['punch_datetime'].dt.strftime('%Y-%m-%d')

    # remove empty emp_id
    work = work[work['emp_id'].str.lower() != 'nan']
    work = work[work['emp_id'].str.strip() != '']

    work = work.sort_values(['emp_id', 'punch_datetime'])

    meta = {
        'detected_columns': detected,
        'total_rows': len(work),
        'employees': work['emp_id'].nunique(),
        'date_range': [str(work['date'].min()), str(work['date'].max())] if len(work)>0 else []
    }
    return work, meta

def process_attendance(
    input_path: str = None,
    df: pd.DataFrame = None,
    company_name: str = "Autocrat Solutions",
    month: Optional[str] = None,
    working_hours_threshold: float = 4.0,  # half day threshold
    full_day_hours: float = 8.0,
    late_threshold: str = "10:00",
    holidays: List[str] = None,
    weekoff: List[int] = None,  # 0=Mon ... 6=Sun, default Sunday
    output_path: str = None,
):
    """
    Main processor
    """
    if holidays is None:
        holidays = []
    if weekoff is None:
        weekoff = [6]  # Sunday

    if df is None:
        if input_path is None:
            raise ValueError("Provide input_path or df")
        # read excel - try all sheets, concat
        xls = pd.ExcelFile(input_path)
        dfs = []
        for sheet in xls.sheet_names:
            try:
                temp = pd.read_excel(input_path, sheet_name=sheet)
                if len(temp) > 0:
                    dfs.append(temp)
            except:
                continue
        if not dfs:
            raise ValueError("No readable sheets found")
        raw = pd.concat(dfs, ignore_index=True) if len(dfs)>1 else dfs[0]
    else:
        raw = df

    cleaned, meta = clean_dataframe(raw)

    if cleaned.empty:
        raise ValueError("No valid attendance records found after cleaning")

    # Parse late threshold
    try:
        late_time = datetime.strptime(late_threshold, "%H:%M").time()
    except:
        late_time = time(10,0)

    # Holiday set
    holiday_dates = set()
    for h in holidays:
        try:
            d = pd.to_datetime(h, dayfirst=True).date()
            holiday_dates.add(d)
        except:
            continue

    # Month filter if provided
    if month:
        try:
            month_dt = pd.to_datetime(month)
            cleaned = cleaned[cleaned['punch_datetime'].dt.to_period('M') == month_dt.to_period('M')]
        except:
            pass

    # Daily aggregation per employee per date
    daily_records = []
    grouped = cleaned.groupby(['emp_id', 'date'])

    for (emp_id, date), group in grouped:
        group = group.sort_values('punch_datetime')
        first_punch = group['punch_datetime'].iloc[0]
        last_punch = group['punch_datetime'].iloc[-1]
        emp_name = group['emp_name'].iloc[0]
        dept = group['department'].iloc[0] if 'department' in group else 'General'

        # calculate hours
        if len(group) == 1:
            # single punch - assume half day? but we count hours 0
            hours = 0.0
        else:
            delta = (last_punch - first_punch).total_seconds() / 3600.0
            # cap unrealistic > 16h to 8h if only 2 punches far apart? keep as is but limit to 24
            hours = min(delta, 24.0)

        # Status logic
        is_holiday = date in holiday_dates
        is_weekoff = date.weekday() in weekoff

        if is_holiday:
            status = 'H'
        elif is_weekoff:
            status = 'WO'
        else:
            if hours >= full_day_hours:
                status = 'P'
            elif hours >= working_hours_threshold:
                status = 'HD'  # Half Day
            elif hours > 0:
                status = 'HD'
            else:
                # single punch still present
                status = 'P' if len(group)>=1 else 'A'

        # Late?
        is_late = first_punch.time() > late_time and status not in ('H','WO')
        # Early exit?
        # ...

        daily_records.append({
            'emp_id': emp_id,
            'emp_name': emp_name,
            'department': dept,
            'date': date,
            'day': date.strftime('%a'),
            'first_punch': first_punch.time().strftime('%H:%M:%S'),
            'last_punch': last_punch.time().strftime('%H:%M:%S'),
            'punch_count': len(group),
            'hours': round(hours,2),
            'status': status,
            'late': 'Yes' if is_late else 'No',
            'is_holiday': is_holiday,
            'is_weekoff': is_weekoff,
        })

    daily_df = pd.DataFrame(daily_records)

    # Now create matrix for all employees and all dates in range
    if daily_df.empty:
        raise ValueError("No daily records aggregated")

    all_dates = sorted(daily_df['date'].unique())
    # if month given, generate full month dates
    if month:
        try:
            mdt = pd.to_datetime(month)
            start = mdt.replace(day=1).date()
            # end of month
            if mdt.month == 12:
                end = mdt.replace(day=31).date()
            else:
                end = (mdt.replace(day=1) + pd.offsets.MonthEnd(0)).date()
            # generate range
            all_dates = [start + timedelta(days=i) for i in range((end-start).days+1)]
        except:
            pass

    employees = daily_df[['emp_id','emp_name','department']].drop_duplicates()

    # Build matrix
    matrix_rows = []
    summary_rows = []
    for _, emp in employees.iterrows():
        emp_id = emp['emp_id']
        emp_daily = daily_df[daily_df['emp_id']==emp_id].set_index('date')

        row = {'emp_id': emp_id, 'emp_name': emp['emp_name'], 'department': emp['department']}
        present = 0
        absent = 0
        half = 0
        holiday = 0
        weekoff_c = 0
        late = 0
        total_hours = 0

        for d in all_dates:
            if d in emp_daily.index:
                rec = emp_daily.loc[d]
                # if duplicate index (shouldn't), take first
                if isinstance(rec, pd.DataFrame):
                    rec = rec.iloc[0]
                status = rec['status']
                row[d.strftime('%Y-%m-%d')] = status
                # counts
                if status == 'P': present+=1
                elif status == 'A': absent+=1
                elif status == 'HD': half+=1
                elif status == 'H': holiday+=1
                elif status == 'WO': weekoff_c+=1
                if rec['late']=='Yes': late+=1
                total_hours+= rec['hours']
            else:
                # No punch -> Absent or WO/H
                if d in holiday_dates:
                    row[d.strftime('%Y-%m-%d')] = 'H'
                    holiday+=1
                elif d.weekday() in weekoff:
                    row[d.strftime('%Y-%m-%d')] = 'WO'
                    weekoff_c+=1
                else:
                    row[d.strftime('%Y-%m-%d')] = 'A'
                    absent+=1

        matrix_rows.append(row)

        summary_rows.append({
            'emp_id': emp_id,
            'emp_name': emp['emp_name'],
            'department': emp['department'],
            'Total Days': len(all_dates),
            'Present (P)': present,
            'Absent (A)': absent,
            'Half Day (HD)': half,
            'Holidays (H)': holiday,
            'Week Off (WO)': weekoff_c,
            'Late Marks': late,
            'Total Hours': round(total_hours,2),
            'Avg Hours/Day': round(total_hours/max(present+half,1),2),
            'Attendance %': round((present + half*0.5)/max(len(all_dates)-holiday-weekoff_c,1)*100,2)
        })

    matrix_df = pd.DataFrame(matrix_rows)
    summary_df = pd.DataFrame(summary_rows)

    # Cleaned logs sheet
    cleaned_export = cleaned.copy()
    cleaned_export['punch_datetime'] = cleaned_export['punch_datetime'].astype(str)

    # Config sheet
    config_data = {
        'Parameter': ['Company Name','Month','Working Hours Threshold','Full Day Hours','Late Threshold','WeekOff (0=Mon)','Holidays','Generated On','Total Employees','Date Range'],
        'Value': [company_name, month or 'Auto', working_hours_threshold, full_day_hours, late_threshold, str(weekoff), ', '.join([str(d) for d in holiday_dates]), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), len(employees), f"{all_dates[0]} to {all_dates[-1]}" if all_dates else ""]
    }
    config_df = pd.DataFrame(config_data)

    # If output path, save with styling
    if output_path:
        save_formatted_excel(output_path, cleaned_export, daily_df, matrix_df, summary_df, config_df, company_name)

    return {
        'cleaned': cleaned_export,
        'daily': daily_df,
        'matrix': matrix_df,
        'summary': summary_df,
        'config': config_df,
        'meta': meta
    }

def save_formatted_excel(output_path, cleaned_df, daily_df, matrix_df, summary_df, config_df, company_name="Autocrat Solutions"):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write sheets
        config_df.to_excel(writer, sheet_name='Config', index=False)
        summary_df.to_excel(writer, sheet_name='Monthly Summary', index=False)
        daily_df.to_excel(writer, sheet_name='Daily Logs', index=False)
        matrix_df.to_excel(writer, sheet_name='Attendance Matrix', index=False)
        cleaned_df.to_excel(writer, sheet_name='Cleaned Raw', index=False)

        # Style each sheet
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        status_colors = {
            'P': PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            'A': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            'HD': PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
            'H': PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
            'WO': PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        }

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # header style
            for col in range(1, ws.max_column+1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            # auto width
            for col in range(1, ws.max_column+1):
                max_len = 0
                col_letter = get_column_letter(col)
                for row in range(1, min(100, ws.max_row+1)):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len+4, 30)

            # special coloring for matrix
            if sheet_name == 'Attendance Matrix':
                for r in range(2, ws.max_row+1):
                    for c in range(4, ws.max_column+1):
                        cell = ws.cell(row=r, column=c)
                        v = str(cell.value).strip()
                        if v in status_colors:
                            cell.fill = status_colors[v]
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border

            # Title row for summary
            if sheet_name == 'Monthly Summary':
                ws.insert_rows(1)
                ws['A1'] = f"{company_name} - Attendance Summary"
                ws['A1'].font = Font(bold=True, size=14, color="0F172A")
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

    print(f"Saved formatted attendance to {output_path}")

# CLI usage
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Autocrat Attendance Processor - Liquid Glass")
    parser.add_argument("input", help="Input Excel file")
    parser.add_argument("-o", "--output", help="Output Excel file", default="attendance_output.xlsx")
    parser.add_argument("-c", "--company", default="Autocrat Solutions")
    parser.add_argument("-m", "--month", help="Month YYYY-MM", default=None)
    parser.add_argument("--late", default="10:00", help="Late threshold HH:MM")
    args = parser.parse_args()
    process_attendance(input_path=args.input, output_path=args.output, company_name=args.company, month=args.month, late_threshold=args.late)
